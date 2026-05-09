"""
Excel Export Engine.
Generates production-ready Excel workbooks with formatting, dropdowns, and validation highlighting.
"""
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from app.models.schemas import ValidationReport, ValidationSeverity
from app.blueprints.topic_master import TopicMasterBlueprint
from app.blueprints.assessment_master import AssessmentMasterBlueprint
from app.blueprints.question_master import QuestionMasterBlueprint
from app.config import OUTPUT_DIR

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
ERROR_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class ExcelExportEngine:
    """Generates formatted Excel workbooks from dump sheet data."""

    def __init__(self, languages: List[str] = None):
        self.languages = languages or ["English"]

    def export(self, sheet_data: Dict[str, Any], output_path: Optional[str] = None,
               validation_report: Optional[ValidationReport] = None) -> str:
        """Export complete dump sheet workbook."""
        if not output_path:
            output_path = str(OUTPUT_DIR / "dump_sheet.xlsx")

        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Create all three sheets
        self._create_sheet(wb, TopicMasterBlueprint.SHEET_NAME,
                          TopicMasterBlueprint.get_columns(self.languages),
                          sheet_data.get("topic_master", []))

        self._create_sheet(wb, AssessmentMasterBlueprint.SHEET_NAME,
                          AssessmentMasterBlueprint.get_columns(),
                          sheet_data.get("assessment_master", []))

        self._create_sheet(wb, QuestionMasterBlueprint.SHEET_NAME,
                          QuestionMasterBlueprint.get_columns(self.languages),
                          sheet_data.get("question_master", []))

        # Apply validation highlighting if report exists
        if validation_report:
            self._apply_error_highlighting(wb, validation_report)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Exported workbook to {output_path}")
        return output_path

    def _create_sheet(self, wb: Workbook, sheet_name: str,
                      columns: List[Dict], data: List[Dict]):
        """Create a formatted sheet with headers, data, and dropdowns."""
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["display"])
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(col_def["display"]) + 4)

            # Add dropdown validations
            if "allowed_values" in col_def and col_def["allowed_values"]:
                values = ",".join(str(v) for v in col_def["allowed_values"])
                dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=not col_def.get("required"))
                dv.error = f"Invalid value for {col_def['display']}"
                dv.errorTitle = "Invalid Input"
                col_letter = get_column_letter(col_idx)
                dv.sqref = f"{col_letter}2:{col_letter}1048576"
                ws.add_data_validation(dv)

        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_def in enumerate(columns, 1):
                val = row_data.get(col_def["name"], "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Freeze header row
        ws.freeze_panes = "A2"
        # Auto-filter
        if columns:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(2, len(data) + 1)}"

    def _apply_error_highlighting(self, wb: Workbook, report: ValidationReport):
        """Apply red/yellow highlighting to cells with validation errors."""
        for error in report.errors:
            if error.sheet not in wb.sheetnames:
                continue
            ws = wb[error.sheet]

            # Find column index
            col_idx = None
            for idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=idx).value
                if header and (header.lower().replace(" ", "_") == error.column or
                               header.lower() == error.column.replace("_", " ")):
                    col_idx = idx
                    break

            if col_idx and error.row + 1 <= ws.max_row:
                cell = ws.cell(row=error.row + 1, column=col_idx)
                if error.severity == ValidationSeverity.CRITICAL:
                    cell.fill = ERROR_FILL
                else:
                    cell.fill = WARNING_FILL
                # Add comment with error message
                from openpyxl.comments import Comment
                cell.comment = Comment(error.message, "Validation Engine")

    def export_validation_report(self, report: ValidationReport,
                                  output_path: Optional[str] = None) -> str:
        """Export validation report as a separate Excel file."""
        if not output_path:
            output_path = str(OUTPUT_DIR / "validation_report.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Report"

        # Summary section
        ws["A1"] = "Validation Summary"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A3"] = "Total Rules Checked:"
        ws["B3"] = report.total_rules_checked
        ws["A4"] = "Total Errors:"
        ws["B4"] = report.total_errors
        ws["A5"] = "Critical Errors:"
        ws["B5"] = report.critical_errors
        ws["A6"] = "Warnings:"
        ws["B6"] = report.warnings
        ws["A7"] = "Valid:"
        ws["B7"] = "Yes" if report.is_valid else "No"

        # Error details
        headers = ["Sheet", "Row", "Column", "Rule", "Severity", "Message", "Current Value"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        for err_idx, err in enumerate(report.errors, 10):
            ws.cell(row=err_idx, column=1, value=err.sheet).border = THIN_BORDER
            ws.cell(row=err_idx, column=2, value=err.row).border = THIN_BORDER
            ws.cell(row=err_idx, column=3, value=err.column).border = THIN_BORDER
            ws.cell(row=err_idx, column=4, value=err.rule).border = THIN_BORDER
            sev_cell = ws.cell(row=err_idx, column=5, value=err.severity.value)
            sev_cell.border = THIN_BORDER
            if err.severity == ValidationSeverity.CRITICAL:
                sev_cell.fill = ERROR_FILL
                sev_cell.font = Font(color="FFFFFF", bold=True)
            else:
                sev_cell.fill = WARNING_FILL
            ws.cell(row=err_idx, column=6, value=err.message).border = THIN_BORDER
            ws.cell(row=err_idx, column=7, value=err.current_value or "").border = THIN_BORDER

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(output_path)
        return output_path
